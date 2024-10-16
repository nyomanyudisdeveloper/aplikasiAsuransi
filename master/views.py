from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.core import serializers
from master.models import Assumption, Pattern, InsuranceContract
import openpyxl
from datetime import date

from django import forms

def pattern(request):
    if request.user.is_authenticated:
        list_table_detail = Pattern.objects.all().values_list("periode","claim_payment_pattern","claim_incurred_pattern","premium_payment_pattern","month_1","month_2","month_3","month_4","month_5","month_6","month_7","month_8","month_9","month_10","month_11","month_12")
        list_table_header = ["Periode (Month)","Claim Payment Pattern","Claim Incurred Pattern","Premium Payment Pattern","1","2","3","4","5","6","7","8","9","10","11","12"]
        context = {
            "list_table_header":list_table_header,
            "list_table_detail":list_table_detail
        }
        return render(request,"master/pattern.html",context)
    else:
        return redirect("/accounts/login")
    
def insurance_contract(request):
    if request.user.is_authenticated:
        list_table_detail = InsuranceContract.objects.all().values_list()
        list_table_header = ["No","REPORTING_DT","ENTITY_ID","INSURANCE_CONTRACT_ID","ENDORSEMENT_TYPE_CD","ISSUE_DT","PRODUCT_LINE_ID","PRODUCT_ID","COHORT_ID","CEDED_FLG","INSURANCE_TYPE","REINS_PROP_COVER_FLG","REINS_TREATY_FLG","ENDORSEMENT_DT","INTERCO_CD","INTERCO_FLG","BRANCH_CD","CHANNEL_CD","APPROACH_CD","TRANSITION_REPORTING_APPROACH_CD","BEGIN_COV_DT","END_COV_DT","CURRENCY_CD"," PREM_AMT","COMMISSION_AMT","PD","LGD","PROFITABILITY","PERIODE_PEMBAYARAN"]
        
        context = {
            "list_table_header":list_table_header,
            "list_table_detail":list_table_detail
        }
        return render(request,"master/insurance_contract.html",context)
    else:
        return redirect("/accounts/login")
    
def assumption(request):
    if request.user.is_authenticated:
        # list_table_detail = serializers.serialize("json", Assumption.objects.all().values('name','ratio'))
        list_table_detail = Assumption.objects.all().values_list("name","ratio")
        # return HttpResponse(list_table_detail)
        list_table_header = ["Assumption","Ratio"]
        context = {
            "list_table_header":list_table_header,
            "list_table_detail":list_table_detail
        }
        return render(request,"master/assumption.html",context)
    else:
        return redirect("/accounts/login")

def assumption_export_template(request):
    # Create a workbook and activate the worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define column headers
    sheet.append(['ASSUMPTION', 'RATIO'])

    # # Fetch data from the model
    data = Assumption.objects.all()

    # # Loop through the data and write it into the Excel sheet
    for obj in data:
        sheet.append([obj.name, obj.ratio])
    
    # Set the HTTP response headers to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=table_assumption_master.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def insurance_contract_export_template(request):
    # Create a workbook and activate the worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define column headers
    sheet.append(["REPORTING_DT","ENTITY_ID","INSURANCE_CONTRACT_ID","ENDORSEMENT_TYPE_CD","ISSUE_DT","PRODUCT_LINE_ID","PRODUCT_ID","COHORT_ID","CEDED_FLG","INSURANCE_TYPE","REINS_PROP_COVER_FLG","REINS_TREATY_FLG","ENDORSEMENT_DT","INTERCO_CD","INTERCO_FLG","BRANCH_CD","CHANNEL_CD","APPROACH_CD","TRANSITION_REPORTING_APPROACH_CD","BEGIN_COV_DT","END_COV_DT","CURRENCY_CD"," PREM_AMT","COMMISSION_AMT","PD","LGD","PROFITABILITY","PERIODE_PEMBAYARAN"])

    # # Fetch data from the model
    data = InsuranceContract.objects.all().values_list()
    print("data = ",data);
    # # Loop through the data and write it into the Excel sheet
    for detail in data:
        list_temp = []
        sheet.append(detail[1:])
    
    # Set the HTTP response headers to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=table_insurance_contract_master.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def pattern_export_template(request):
    # Create a workbook and activate the worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define column headers
    sheet.append(['Periode (Month)', 'Claim Payment Pattern','Claim Incurred Pattern','Premium Payment Pattern','1','2','3','4','5','6','7','8','9','10','11','12'])

    # # Fetch data from the model
    data = Pattern.objects.all()

    # # Loop through the data and write it into the Excel sheet
    for obj in data:
        sheet.append([obj.periode, obj.claim_payment_pattern,obj.claim_incurred_pattern,obj.premium_payment_pattern,obj.month_1,obj.month_2,obj.month_3,obj.month_4,obj.month_5,obj.month_6,obj.month_7,obj.month_8,obj.month_9,obj.month_10,obj.month_11,obj.month_12])
    
    # Set the HTTP response headers to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=table_pattern_master.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def assumption_import_data_from_template(request):
    try:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]

        Assumption.objects.all().delete()

        isHeader = True
        for row in worksheet.iter_rows():
            if not isHeader:
                row_data = []
                for cell in row:
                    row_data.append(str(cell.value))
                assumption = Assumption(name=row_data[0],ratio=row_data[1])
                assumption.save()
            else:
                isHeader = False

        return redirect("/master/assumption/")
    except:
        context = {
            "error":"Something wrong with import excel, Please check again its format"
        }
        return render(request,"master/pattern.html",context)
    
def insurance_contract_import_template(request):
    # try:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]

        InsuranceContract.objects.all().delete()

        isHeader = True
        for row in worksheet.iter_rows():
            if not isHeader:
                insurance_contract = InsuranceContract()
                for index,cell in enumerate(row):
                    if str(cell.value) != "None":
                        if index == 0:
                            insurance_contract.reporting_date = str(cell.value)[0:10]
                        elif index == 1:
                            insurance_contract.entity_id = int(cell.value)
                        elif index == 2:
                            insurance_contract.insurance_contract_id = str(cell.value)
                        elif index == 3:
                            insurance_contract.endorsement_type_cd = int(cell.value)
                        elif index == 4:
                            insurance_contract.issue_dt = str(cell.value)[0:10]
                        elif index == 5:
                            insurance_contract.product_line_id = str(cell.value)
                        elif index == 6:
                            insurance_contract.product_id = str(cell.value)
                        elif index == 7:
                            insurance_contract.cohort_id = str(cell.value)
                        elif index == 8:
                            insurance_contract.cedeg_flg = str(cell.value)
                        elif index == 9:
                            insurance_contract.insurance_type = str(cell.value)
                        elif index == 10:
                            insurance_contract.reins_prop_cover_flg = str(cell.value)
                        elif index == 11:
                            insurance_contract.reins_treaty_flg = str(cell.value)
                        elif index == 12:
                            insurance_contract.endorsement_dt = str(cell.value)[0:10]
                        elif index == 13: 
                            insurance_contract.interco_cd = str(cell.value)
                        elif index == 14:
                            insurance_contract.interco_flg = str(cell.value)
                        elif index == 15:
                            insurance_contract.branch_cd = str(cell.value)
                        elif index == 16:
                            insurance_contract.channel_cd = str(cell.value)
                        elif index == 17:
                            insurance_contract.approach_cd = str(cell.value)
                        elif index == 18:
                            insurance_contract.transition_reporting_approach_cd = str(cell.value)
                        elif index == 19:
                            insurance_contract.begin_cov_dt = str(cell.value)[0:10]
                        elif index == 20:
                            insurance_contract.end_cov_dt = str(cell.value)[0:10]
                        elif index == 21:
                            insurance_contract.currency_cd = str(cell.value)
                        elif index == 22:
                            insurance_contract.prem_amt = int(cell.value)
                        elif index == 23:
                            insurance_contract.commision_amt = int(cell.value)
                        elif index == 24:
                            insurance_contract.pd = int(cell.value)
                        elif index == 25:
                            insurance_contract.lgd = int(cell.value)
                        elif index == 26:
                            insurance_contract.profitability = str(cell.value)
                        elif index == 27:
                            insurance_contract.periode_pembayaran = int(cell.value)
                insurance_contract.save()
            else:
                isHeader = False

        return redirect("/master/insurance_contract/")
    # except:
    #     context = {
    #         "error":"Something wrong with import excel, Please check again its format"
    #     }
    #     return render(request,"master/insurance_contract.html",context)
    

def pattern_import_data_from_template(request):
    try:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file,data_only=True)
        sheets = wb.sheetnames
        worksheet = wb[sheets[0]]

        Pattern.objects.all().delete()

        isHeader = True
        for row in worksheet.iter_rows():
            if not isHeader:
                row_data = []
                for index,cell in enumerate(row):
                    if index == 0:
                        row_data.append(int(cell.value))
                    elif str(cell.value) == "None":
                        row_data.append(0.0)
                    else:
                        row_data.append(float(cell.value))
                pattern = Pattern(month_1=row_data[4])
                pattern.periode = row_data[0]
                pattern.claim_payment_pattern = row_data[1]
                pattern.claim_incurred_pattern = row_data[2]
                pattern.premium_payment_pattern = row_data[3]
                pattern.month_1 = row_data[4]
                pattern.month_2 = row_data[5]
                pattern.month_3 = row_data[6]
                pattern.month_4 = row_data[7]
                pattern.month_5 = row_data[8]
                pattern.month_6 = row_data[9]
                pattern.month_7 = row_data[10]
                pattern.month_8 = row_data[11]
                pattern.month_9 = row_data[12]
                pattern.month_10 = row_data[13]
                pattern.month_11 = row_data[14]
                pattern.month_12 = row_data[15]
                pattern.save()
            else:
                isHeader = False

        return redirect("/master/pattern")
    except:
        context = {
            "error":"Something wrong with import excel, Please check again its format"
        }
        return render(request,"master/pattern.html",context)
